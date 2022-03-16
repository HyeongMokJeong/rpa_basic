import smtplib
from imap_tools import MailBox
from account import *
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook() # 새 워크북 생성
ws = wb.active # 현재 활성화된 sheet 가져옴

title = ("순번", "닉네임", "전화번호")
ws.append(title)

list_address = []
list = []

# 당첨자 선정
with MailBox("imap.gmail.com", 993).login(EMAIL_ADDRESS, EMAIL_PASSWORD, initial_folder="INBOX") as mailbox:
    for msg in mailbox.fetch('(ON 08-Jul-2021)'):
        if "파이썬 특강 신청합니다" in msg.subject: 
            if "/" in msg.text:
                # 제목 맞고 내용에 /가 들어가 있다면
                # 리스트에 보낸사람과 내용 저장(내용을 spilt으로 닉네임/번호 구분할 예정)
                list_address.append(msg.from_)
                list.append(msg.text)

# 메일 발송
for i in range(1,len(list)+1):
# 선정 안내 메일 양식
    info = list[i-1].split("/")
    # info[0]이 닉네임, info[1]이 전화번호
    if i < 4:
        msg = EmailMessage()
        msg["Subject"] = "파이썬 특강 안내 [선정]" # 제목
        msg["From"] = EMAIL_ADDRESS # 보내는 사람
        msg["To"] = f"{list_address[i-1]}" # 받는 사람
        msg.set_content(f"{info[0]}님 축하드립니다. 특강 대상자로 선정되셨습니다.  (선정순번 {i}번)") # 본문 내용

        # 메일 발송
        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)
        
        # 엑셀 파일에 저장
        ws[f"A{i+1}"] = i # 순번
        ws[f"B{i+1}"] = info[0]
        ws[f"C{i+1}"] = info[1]

    # 탈락자일 경우
    else:
        msg = EmailMessage()
        msg["Subject"] = "파이썬 특강 안내 [탈락]" # 제목
        msg["From"] = EMAIL_ADDRESS # 보내는 사람
        msg["To"] = f"{list_address[i-1]}" # 받는 사람
        msg.set_content(f"{info[0]}님 아쉽게도 탈락입니다. 취소 인원이 발생하는 경우 연락드리겠습니다.  (대기순번 {i-3}번)") # 본문 내용

        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)

# 전체 셀 가운데 정렬
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save("[선정 명단 엑셀].xlsx")
