from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.title = "Mysheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB값 입력하여 탭 색상 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 인덱스에 sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 접근

print(wb.sheetnames) # 모든 sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test" # NewSheet의 A1 셀에 입력
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")