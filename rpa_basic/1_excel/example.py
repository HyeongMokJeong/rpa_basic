from openpyxl import load_workbook
from openpyxl.styles import Alignment

wb = load_workbook("scores.xlsx")
ws = wb.active

# 퀴즈2 점수를 10으로 수정
col_q2 = ws["D"] # D 행의 모든 데이터 저장
for col in col_q2:
    if isinstance(col.value, int):
        col.value = 10

# 타이틀 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

# 총점 정보
for i in range(1,11):
    result = ws[f"H{i+1}"]
    result.value = f"=SUM(B{i+1}:G{i+1})"  

# 성적 정보
for row in ws.iter_rows(max_row=ws.max_row, min_col=2, max_col=7):
    for a in range(0, 6):
        if type(row[a].value) == int:
            result = 0
            result += row[a].value
    print(result)