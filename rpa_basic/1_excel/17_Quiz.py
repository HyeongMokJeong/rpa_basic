from openpyxl import workbook
from openpyxl.styles import Alignment
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active

title = ("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트")
scores = [(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)]

# 최종 성적 데이터 넣기
ws.append(title)
for s in scores:
    ws.append(s)

# 퀴즈2 점수를 10으로 수정
col_q2 = ws["D"] # D 행의 모든 데이터 저장
for col in col_q2:
    if isinstance(col.value, int):
        col.value = 10

# 타이틀 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

# 총점 정보
# for i in range(1,11):
#     result = ws[f"H{i+1}"]
#     result.value = f"=SUM(B{i+1}:G{i+1})" 
for idx, score in enumerate(scores, start=2): #idx가 2부터 시작
    result = sum(score[1:]) + 10 - score[3]
    ws[f"H{idx}"] = f"=SUM(B{idx}:G{idx})"

# 성적 정보 
    if result >= 90:
        ws[f"I{idx}"] = "A"
    elif 90 > result >= 80:
        ws[f"I{idx}"] = "B"
    elif 80 > result >= 70:
        ws[f"I{idx}"] = "C"
    else:
        ws[f"I{idx}"] = "D"
# 출석 5 미만은 F, 글자 빨간색
    if score[1] < 5:
        ws[f"I{idx}"] = "F"
        ws[f"I{idx}"].font = Font(color="FF0000")

# 전체 셀 정렬
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 틀 고정
ws.freeze_panes = "A2"

wb.save("scores.xlsx")